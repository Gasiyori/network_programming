from openpyxl import Workbook
wb = Workbook()

ws = wb.active # 활성화된 워크시트 선택
ws['A1'] = 42 # 셀에 데이터 추가
ws.append([1, 2, 3]) # 그 다음 행(A2, B2, C2)에 데이터 추가

import datetime
ws['A2'] = datetime.datetime.now() # A2행의 1을 날짜로 교체
wb.save("sample.xlsx") # 파일명